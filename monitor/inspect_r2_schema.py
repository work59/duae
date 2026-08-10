import logging
import openpyxl
import io
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional, Tuple


from monitor_r2 import (
    partition_date_for_listing,
)

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("monitor")
#───────────────────────────────────────────────────────────────────────────────

def list_excel_files(client, bucket: str, prefix: str) -> List[Dict]:
    """
    Return all .xlsx objects under *prefix*.
    Each item: {key, size, last_modified}
    Handles pagination automatically.
    """
    results = []
    paginator = client.get_paginator("list_objects_v2")
    for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
        for obj in page.get("Contents", []):
            if obj["Key"].endswith(".xlsx"):
                results.append(
                    {
                        "key":           obj["Key"],
                        "size_bytes":    obj["Size"],
                        "last_modified": obj["LastModified"].isoformat(),
                    }
                )
    return results

def download_excel(client, bucket: str, key: str) -> Optional[bytes]:
    """Download an .xlsx file from R2 and return raw bytes."""
    try:
        resp = client.get_object(Bucket=bucket, Key=key)
        return resp["Body"].read()
    except Exception as exc:
        log.warning(f"Could not download {key}: {exc}")
        return None


def inspect_excel(raw: bytes, file_key: str) -> Dict:
    """
    Open an xlsx from bytes and return:
      sheets: [{name, row_count, columns: [str]}]
      readable: bool
      error: str | None
    """
    result = {"file_key": file_key, "readable": False, "sheets": [], "error": None}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Read header row
            headers = []
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if header_row:
                headers = [str(c) for c in header_row if c is not None]
            # Count data rows (fast — don't load everything into memory)
            row_count = sum(1 for _ in rows_iter)
            result["sheets"].append(
                {"name": sheet_name, "row_count": row_count, "columns": headers}
            )
        wb.close()
        result["readable"] = True
    except Exception as exc:
        result["error"] = str(exc)
        log.warning(f"Error inspecting {file_key}: {exc}")
    return result

def accumulate_stats(
    existing: Dict,
    scraper_name: str,
    inspected: Dict,
    size_bytes: int,
    run_date: str,
) -> None:
    """
    Merge new observations into the *existing* stats dict (in-place).
    Structure:
      scraper_name:
        observed_dates: [...]
        file_size_kb: {min, max, last}
        sheets:
          sheet_name:
            row_count: {min, max, last}
            columns: [...]
    """
    entry = existing.setdefault(scraper_name, {
        "observed_dates":  [],
        "file_size_kb":    {"min": None, "max": None, "last": None},
        "sheets":          {},
    })

    if run_date not in entry["observed_dates"]:
        entry["observed_dates"].append(run_date)
        entry["observed_dates"] = sorted(entry["observed_dates"])[-30:]  # keep last 30

    kb = round(size_bytes / 1024, 1)
    fs = entry["file_size_kb"]
    fs["last"] = kb
    fs["min"]  = min(kb, fs["min"]) if fs["min"] is not None else kb
    fs["max"]  = max(kb, fs["max"]) if fs["max"] is not None else kb

    for sheet in inspected.get("sheets", []):
        sname = sheet["name"]
        se    = entry["sheets"].setdefault(sname, {
            "row_count": {"min": None, "max": None, "last": None},
            "columns":   [],
        })
        rc = sheet["row_count"]
        se["row_count"]["last"] = rc
        se["row_count"]["min"]  = min(rc, se["row_count"]["min"]) if se["row_count"]["min"] is not None else rc
        se["row_count"]["max"]  = max(rc, se["row_count"]["max"]) if se["row_count"]["max"] is not None else rc
        # Merge new columns (union)
        new_cols = [c for c in sheet["columns"] if c not in se["columns"]]
        se["columns"].extend(new_cols)

def r2_base_prefix(r2_path_raw: str) -> Tuple[str, Optional[str]]:
    """
    Convert config r2_path like '{r2_bucket}/DKSA/Vehicles' into (base, category).

    The category is a subfolder that sits UNDER the date partition
    (DKSA/year=.../month=.../day=.../Vehicles/excel/...), so callers need
    both pieces to build a correct, scraper-specific prefix.

        '{r2_bucket}/DKSA/Vehicles' -> ('DKSA', 'Vehicles')
        '{r2_bucket}/DKSA'          -> ('DKSA', None)
    """
    path = r2_path_raw.strip()
    if path.startswith("{"):
        path = path.split("/", 1)[1] if "/" in path else path

    parts = path.strip("/").split("/")
    if not parts or not parts[0]:
        return path.strip("/"), None

    base = parts[0]
    category = "/".join(parts[1:]) if len(parts) > 1 else None
    return base, category

def partition_date_for_data_date(dt: datetime) -> datetime:
    """
    Return the partition date for the given data date.
    
    For DKSA, the data is stored under:
    DKSA/year=YYYY/month=MM/day=DD/{Category}/excel/
    
    So we use the data date directly.
    """
    return dt


def excel_prefixes_for_date(base: str, category: Optional[str], dt: datetime) -> List[str]:
    """
    Build the R2 date-partition prefix for Excel discovery, scoped to one
    scraper's own category folder.

    Actual structure:
    DKSA/year=2026/month=08/day=02/Vehicles/excel/Audi.xlsx
    """
    base = base.strip("/")
    date_part = f"year={dt.year}/month={dt.month:02d}/day={dt.day:02d}"

    if category:
        prefix = f"{base}/{date_part}/{category.strip('/')}/excel/"
    else:
        # No category known — fall back to the broad date folder (legacy behavior).
        prefix = f"{base}/{date_part}/"

    return [prefix]